# -*- coding: utf-8 -*-
"""
③ XBRL Financial Analyzer — TDnet XBRL取得 & 財務分析ツール

【機能概要】
1. TDnetの一覧ページからXBRLファイル（ZIP）をダウンロード
2. XBRLを解析し、財務諸表データ（P/L, B/S, CF）を抽出
3. Excelファイルに出力（スプレッドシートで操作可能な形式）
4. 財務分析を実行:
   - 前期比の増減率計算
   - 大きく変化した勘定科目の検出
   - 売上高利益率（営業利益率、経常利益率、純利益率）の計算

【実行例】
  python "③_xbrl_financial_analyzer.py" --target "20260202"
  python "③_xbrl_financial_analyzer.py" --target "20260202" --code 7203
  python "③_xbrl_financial_analyzer.py" --target "202602" --threshold 0.15

【前提】
  pip install requests beautifulsoup4 lxml pandas openpyxl
"""

import os
import sys
import datetime
import requests
import pandas as pd
import time
import re
import unicodedata
import argparse
import zipfile
import io
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from pathlib import Path
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Windows cp932 で絵文字が出力できない問題の回避
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass


# ============================================================
# 定数・設定
# ============================================================

DEFAULT_TARGET_SPEC = "20260203"

# 保存先（Windowsローカル）
DEFAULT_SAVE_ROOT = r"G:\マイドライブ\TDnet_XBRL"

# TDnet負荷軽減
PAGE_SLEEP_SEC = 3
XBRL_SLEEP_SEC = 1

# 変化率の閾値（デフォルト20%以上の変化を「大きな変動」とする）
DEFAULT_CHANGE_THRESHOLD = 0.20

# 除外キーワード（タイトルに含まれたら完全除外）
EXCLUDE_KEYWORDS = ["ＥＴＦ", "ETF", "ETN", "ＥＴＮ", "_MAXIS"]

# TDnet URL template
BASE_URL_TEMPLATE = "https://www.release.tdnet.info/inbs/I_list_{}_{}.html"

# HTTP設定
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "Referer": "https://www.release.tdnet.info/index.html",
}
COOKIES = {"cb_agree": "0"}


# ============================================================
# XBRLラベルマッピング（要素名 → 日本語名称）
# ============================================================

XBRL_LABEL_MAP = {
    # --- 損益計算書 (P/L) ---
    "NetSales": "売上高",
    "Revenue": "売上収益（IFRS）",
    "OperatingRevenue1": "営業収益",
    "CostOfSales": "売上原価",
    "GrossProfit": "売上総利益",
    "SellingGeneralAndAdministrativeExpenses": "販売費及び一般管理費",
    "OperatingIncome": "営業利益",
    "NonOperatingIncome": "営業外収益",
    "NonOperatingExpenses": "営業外費用",
    "OrdinaryIncome": "経常利益",
    "ExtraordinaryIncome": "特別利益",
    "ExtraordinaryLoss": "特別損失",
    "IncomeBeforeIncomeTaxes": "税引前当期純利益",
    "IncomeTaxes": "法人税等合計",
    "ProfitLoss": "当期純利益",
    "ProfitLossAttributableToOwnersOfParent": "親会社株主に帰属する当期純利益",
    "ComprehensiveIncome": "包括利益",

    # --- 貸借対照表 (B/S) ---
    "CurrentAssets": "流動資産合計",
    "NoncurrentAssets": "固定資産合計",
    "DeferredAssets": "繰延資産",
    "TotalAssets": "総資産",
    "CurrentLiabilities": "流動負債合計",
    "NoncurrentLiabilities": "固定負債合計",
    "TotalLiabilities": "負債合計",
    "NetAssets": "純資産合計",
    "ShareholdersEquity": "株主資本合計",
    "CapitalStock": "資本金",
    "CapitalSurplus": "資本剰余金",
    "RetainedEarnings": "利益剰余金",
    "TreasuryStock": "自己株式",

    # --- キャッシュフロー計算書 (CF) ---
    "NetCashProvidedByUsedInOperatingActivities": "営業活動によるCF",
    "NetCashProvidedByUsedInInvestingActivities": "投資活動によるCF",
    "NetCashProvidedByUsedInFinancingActivities": "財務活動によるCF",
    "CashAndCashEquivalents": "現金及び現金同等物期末残高",
    "IncreaseDecreaseInCashAndCashEquivalents": "現金及び現金同等物の増減額",

    # --- 1株当たり情報 ---
    "EarningsPerShare": "1株当たり当期純利益",
    "DilutedEarningsPerShare": "潜在株式調整後EPS",
    "DividendPerShare": "1株当たり配当額",
    "NetAssetsPerShare": "1株当たり純資産",

    # --- 経営指標 ---
    "EquityToAssetRatio": "自己資本比率（%）",
    "RateOfReturnOnEquity": "自己資本利益率ROE（%）",
    "PriceEarningsRatio": "株価収益率PER（倍）",

    # --- 会社情報 (DEI) ---
    "FilerNameInJapaneseDEI": "提出者名",
    "SecurityCodeDEI": "証券コード",
    "AccountingStandardsDEI": "会計基準",
    "CurrentFiscalYearStartDateDEI": "当期開始日",
    "CurrentFiscalYearEndDateDEI": "当期終了日",
    "CurrentPeriodEndDateDEI": "当四半期末日",
    "TypeOfCurrentPeriodDEI": "当四半期会計期間の種類",
}


# ============================================================
# TDnetサマリー要素名 → 標準要素名マッピング
# TDnetの決算短信サマリーは tse-ed-t 名前空間独自の要素名を使う
# ============================================================

TSE_ELEMENT_MAP = {
    # --- P/L 日本基準 ---
    "NetSales": "NetSales",
    "OperatingIncome": "OperatingIncome",
    "OrdinaryIncome": "OrdinaryIncome",
    "ProfitLoss": "ProfitLoss",
    "ProfitLossAttributableToOwnersOfParent": "ProfitLossAttributableToOwnersOfParent",
    "ComprehensiveIncome": "ComprehensiveIncome",
    # --- P/L IFRS ---
    "SalesIFRS": "NetSales",
    "RevenueIFRS": "NetSales",
    "OperatingIncomeIFRS": "OperatingIncome",
    "ProfitBeforeTaxIFRS": "IncomeBeforeIncomeTaxes",
    "ProfitLossIFRS": "ProfitLoss",
    "ProfitLossAttributableToOwnersOfParentIFRS": "ProfitLossAttributableToOwnersOfParent",
    "ComprehensiveIncomeIFRS": "ComprehensiveIncome",
    # --- 変動率 ---
    "ChangeInNetSales": "ChangeInNetSales",
    "ChangeInOperatingIncome": "ChangeInOperatingIncome",
    "ChangeInOrdinaryIncome": "ChangeInOrdinaryIncome",
    "ChangeInProfitLoss": "ChangeInProfitLoss",
    "ChangeInSalesIFRS": "ChangeInNetSales",
    "ChangeInOperatingIncomeIFRS": "ChangeInOperatingIncome",
    "ChangeInProfitBeforeTaxIFRS": "ChangeInIncomeBeforeTaxes",
    "ChangeInProfitLossIFRS": "ChangeInProfitLoss",
    # --- EPS ---
    "EarningsPerShare": "EarningsPerShare",
    "DilutedEarningsPerShare": "DilutedEarningsPerShare",
    "EarningsPerShareIFRS": "EarningsPerShare",
    "DilutedEarningsPerShareIFRS": "DilutedEarningsPerShare",
    # --- B/S ---
    "TotalAssets": "TotalAssets",
    "NetAssets": "NetAssets",
    "Equity": "ShareholdersEquity",
    "TotalAssetsIFRS": "TotalAssets",
    "NetAssetsIFRS": "NetAssets",
    "EquityIFRS": "ShareholdersEquity",
    "EquityToAssetRatio": "EquityToAssetRatio",
    "EquityToAssetRatioIFRS": "EquityToAssetRatio",
    "BookValuePerShare": "NetAssetsPerShare",
    "BookValuePerShareIFRS": "NetAssetsPerShare",
    # --- 配当 ---
    "DividendPerShare": "DividendPerShare",
    "AnnualDividendPerShare": "DividendPerShare",
    "DividendPerShareIFRS": "DividendPerShare",
    # --- CF ---
    "CashFlowsFromOperatingActivities": "NetCashProvidedByUsedInOperatingActivities",
    "CashFlowsFromInvestingActivities": "NetCashProvidedByUsedInInvestingActivities",
    "CashFlowsFromFinancingActivities": "NetCashProvidedByUsedInFinancingActivities",
    "CashAndEquivalents": "CashAndCashEquivalents",
    "CashFlowsFromOperatingActivitiesIFRS": "NetCashProvidedByUsedInOperatingActivities",
    "CashFlowsFromInvestingActivitiesIFRS": "NetCashProvidedByUsedInInvestingActivities",
    "CashFlowsFromFinancingActivitiesIFRS": "NetCashProvidedByUsedInFinancingActivities",
    "CashAndEquivalentsIFRS": "CashAndCashEquivalents",
    # --- IFRS追加要素 ---
    "ProfitIFRS": "ProfitLoss",
    "ProfitAttributableToOwnersOfParentIFRS": "ProfitLossAttributableToOwnersOfParent",
    "TotalComprehensiveIncomeIFRS": "ComprehensiveIncome",
    "BasicEarningsPerShareIFRS": "EarningsPerShare",
    "TotalEquityIFRS": "NetAssets",
    "EquityAttributableToOwnersOfParentIFRS": "ShareholdersEquity",
    "EquityAttributableToOwnersOfParentToTotalAssetsRatioIFRS": "EquityToAssetRatio",
}


# ============================================================
# ユーティリティ（①と共通）
# ============================================================

def nfkc(s: str) -> str:
    """Unicode正規化（NFKC）"""
    return unicodedata.normalize("NFKC", str(s))


def safe_filename(s: str, max_len: int = 120) -> str:
    """ファイル名に使える安全な文字列に変換"""
    s = nfkc(s)
    s = re.sub(r'[\\/:*?"<>|]', "_", s)
    s = re.sub(r"\s+", " ", s).strip()
    if len(s) > max_len:
        s = s[:max_len].rstrip()
    return s


def is_excluded(title: str) -> bool:
    """除外キーワードに該当するか判定"""
    if not EXCLUDE_KEYWORDS:
        return False
    t = nfkc(title)
    return any(nfkc(k) in t for k in EXCLUDE_KEYWORDS)


def parse_target_spec(spec: str):
    """日付指定のパース（①と同一仕様）"""
    spec = spec.strip()
    parts = spec.split()

    if len(parts) == 1:
        s = parts[0]
        if re.fullmatch(r"\d{8}", s):
            return s, s, s, "day"
        if re.fullmatch(r"\d{6}", s):
            y = int(s[:4]); m = int(s[4:6])
            start = datetime.date(y, m, 1)
            if m == 12:
                end = datetime.date(y + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                end = datetime.date(y, m + 1, 1) - datetime.timedelta(days=1)
            return start.strftime("%Y%m%d"), end.strftime("%Y%m%d"), s, "month"
        raise ValueError("TARGET_SPEC は 'YYYYMMDD' / 'YYYYMM' / 'YYYYMMDD YYYYMMDD' のいずれかです。")

    if len(parts) == 2:
        d1, d2 = parts
        if not (re.fullmatch(r"\d{8}", d1) and re.fullmatch(r"\d{8}", d2)):
            raise ValueError("範囲指定は 'YYYYMMDD YYYYMMDD' 形式で指定してください。")
        if d1 > d2:
            d1, d2 = d2, d1
        return d1, d2, f"{d1}_{d2}", "range"

    raise ValueError("TARGET_SPEC の指定が不正です。")


def iter_dates_yyyymmdd(d_from: str, d_to: str):
    """YYYYMMDDの範囲で日付を列挙（両端含む）"""
    start = datetime.datetime.strptime(d_from, "%Y%m%d").date()
    end = datetime.datetime.strptime(d_to, "%Y%m%d").date()
    cur = start
    while cur <= end:
        yield cur.strftime("%Y%m%d")
        cur += datetime.timedelta(days=1)


# ============================================================
# Section 1: TDnet XBRL ダウンロード
# ============================================================

def find_xbrl_links(session, target_date_str, code_filter=None):
    """
    TDnetの一覧ページからXBRLリンク（.zip）を取得する。

    Returns:
        list of dict: [{time, code, name, title, xbrl_url}, ...]
    """
    results = []
    page_num = 1

    while True:
        page_str = f"{page_num:03}"
        target_url = BASE_URL_TEMPLATE.format(page_str, target_date_str)

        print(f"   ...Page {page_str} を確認中")
        try:
            res = session.get(target_url, headers=HEADERS, cookies=COOKIES, timeout=60)
        except requests.RequestException as e:
            print(f"   ❌ アクセスエラー: {e}")
            break

        res.encoding = "utf-8"

        # データ無し判定
        if res.status_code == 404 or "該当するデータはありません" in res.text:
            if page_num == 1:
                print("   ⚠️ 該当データなし（休日等の可能性）")
            break

        soup = BeautifulSoup(res.text, "html.parser")
        rows = soup.find_all("tr")

        if len(rows) < 5:
            break

        found_in_page = 0
        for row in rows:
            cols = row.find_all("td")
            if len(cols) < 5:
                continue

            r_time = nfkc(cols[0].get_text(strip=True))
            r_code = nfkc(cols[1].get_text(strip=True))
            r_name = nfkc(cols[2].get_text(strip=True))
            r_title = nfkc(cols[3].get_text(strip=True))

            # 除外
            if is_excluded(r_title):
                continue

            # コードフィルタ
            code4 = (r_code[:4] or "").strip()
            if code_filter and code4 != str(code_filter):
                continue

            # XBRLリンク探索: 全カラムから .zip リンクを探す
            xbrl_url = None
            for col in cols:
                for a_tag in col.find_all("a", href=True):
                    href = a_tag.get("href", "")
                    if href.lower().endswith(".zip"):
                        xbrl_url = urljoin(target_url, href)
                        break
                if xbrl_url:
                    break

            if xbrl_url:
                results.append({
                    "time": r_time,
                    "code": code4,
                    "name": r_name,
                    "title": r_title,
                    "xbrl_url": xbrl_url,
                })
                found_in_page += 1

        page_num += 1
        if PAGE_SLEEP_SEC > 0:
            time.sleep(PAGE_SLEEP_SEC)

    return results


def download_xbrl_zip(session, url, save_path):
    """XBRLのZIPファイルをダウンロード"""
    try:
        r = session.get(url, headers=HEADERS, cookies=COOKIES, stream=True, timeout=60)
        r.raise_for_status()
        with open(save_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=1024 * 256):
                if chunk:
                    f.write(chunk)
        return True
    except Exception as e:
        print(f"   ❌ XBRLダウンロード失敗: {e}")
        return False


# ============================================================
# Section 2: XBRL 解析
# ============================================================

def find_xbrl_instance_in_zip(zip_path):
    """
    ZIPファイルからXBRLインスタンスドキュメントを探す。

    TDnet XBRL の ZIP構造:
      XBRLData/Summary/   *-ixbrl.htm    ← サマリー（決算短信1ページ目）
      XBRLData/Attachment/ *-ixbrl.htm    ← 詳細（財務諸表: B/S, P/L, CF等）
                           *-def.xml      ← 定義（←これはスキップ）
                           *-pre.xml      ← 表示（←これもスキップ）
                           *-cal.xml      ← 計算（←これもスキップ）
                           *-lab.xml      ← ラベル（←これもスキップ）

    優先順位:
      1. Attachment 配下の -ixbrl.htm（最もサイズが大きいもの = B/S全体 等）
      2. Summary 配下の -ixbrl.htm（サマリー情報）
      3. .xbrl ファイル
    """
    with zipfile.ZipFile(zip_path, 'r') as zf:
        ixbrl_attachment = []
        ixbrl_summary = []
        xbrl_files = []

        for name in zf.namelist():
            lower = name.lower()
            if lower.endswith('/'):
                continue

            info = zf.getinfo(name)

            # Inline XBRL（メインデータ）
            if lower.endswith('-ixbrl.htm') or lower.endswith('-ixbrl.html'):
                if 'attachment' in lower:
                    ixbrl_attachment.append((name, info.file_size))
                elif 'summary' in lower:
                    ixbrl_summary.append((name, info.file_size))
                else:
                    ixbrl_attachment.append((name, info.file_size))

            # 通常の XBRL インスタンス
            elif lower.endswith('.xbrl'):
                xbrl_files.append((name, info.file_size))

        # サマリーの iXBRL を返す（決算概要データ）
        # → 最もサイズが大きいファイルを選択
        if ixbrl_summary:
            ixbrl_summary.sort(key=lambda x: -x[1])
            best = ixbrl_summary[0][0]
            content = zf.read(best)
            return best, content

        # Attachment の iXBRL（詳細財務諸表）
        if ixbrl_attachment:
            ixbrl_attachment.sort(key=lambda x: -x[1])
            best = ixbrl_attachment[0][0]
            content = zf.read(best)
            return best, content

        # 通常の XBRL
        if xbrl_files:
            xbrl_files.sort(key=lambda x: -x[1])
            best = xbrl_files[0][0]
            content = zf.read(best)
            return best, content

        print("   ⚠️ XBRLインスタンスが見つかりません")
        return None, None


def parse_contexts(tree):
    """
    XBRLコンテキスト要素を解析して辞書で返す。
    HTMLパーサー使用時はタグ名が小文字化される（xbrli:context 等）。
    """
    contexts = {}

    for elem in tree.iter():
        tag_str = str(elem.tag)
        # 名前空間付き or HTMLパーサーで小文字化された形式
        if '}' in tag_str:
            local_name = etree.QName(tag_str).localname
        else:
            local_name = tag_str
            # HTMLパーサーの場合: "xbrli:context" → "context" 部分を取得
            if ':' in local_name:
                local_name = local_name.split(':', 1)[1]

        if local_name.lower() == "context":
            ctx_id = elem.get("id", "")
            if ctx_id:
                period_info = {}
                for child in elem.iter():
                    child_tag = str(child.tag)
                    if '}' in child_tag:
                        child_name = etree.QName(child_tag).localname
                    else:
                        child_name = child_tag
                        if ':' in child_name:
                            child_name = child_name.split(':', 1)[1]

                    child_lower = child_name.lower()
                    if child_lower in ("startdate", "enddate", "instant"):
                        if child.text:
                            period_info[child_lower] = child.text
                contexts[ctx_id] = period_info

    return contexts


def classify_period(context_ref: str, contexts: dict) -> str:
    """
    コンテキストIDから期間タイプを分類する。

    TDnet XBRL の典型的なコンテキストID例:
      通常:
        CurrentYearDuration        → 当期
        PriorYearDuration          → 前期
        CurrentYearInstant         → 当期末
        PriorYearInstant           → 前期末
      TDnetサマリー:
        CurrentAccumulatedQ3Duration_ConsolidatedMember_ResultMember → 当期
        PriorAccumulatedQ3Duration_ConsolidatedMember_ResultMember   → 前期
        CurrentAccumulatedQ3Instant                                  → 当期末
        PriorAccumulatedQ3Instant                                    → 前期末
        NextAccumulatedFYDuration_ConsolidatedMember_ForecastMember  → 予想
    """
    cr = context_ref.lower()

    # 予想
    if "forecast" in cr or "nextaccumulated" in cr:
        return "予想"

    # 前期
    if cr.startswith("prior") or "prioryear" in cr or "prior1year" in cr or "prioraccumulated" in cr:
        if "instant" in cr:
            return "前期末"
        return "前期"

    # 前四半期
    if "priorquarter" in cr or "prior1quarter" in cr:
        return "前四半期"

    # 当期
    if cr.startswith("current") or "currentyear" in cr or "currentaccumulated" in cr:
        if "instant" in cr:
            return "当期末"
        return "当期"

    # フォールバック
    if "prior" in cr:
        if "instant" in cr:
            return "前期末"
        return "前期"
    if "current" in cr:
        if "instant" in cr:
            return "当期末"
        return "当期"

    return context_ref


def parse_xbrl_content(content: bytes, filename: str):
    """
    XBRLコンテンツを解析し、財務データを抽出する。
    inline XBRL (iXBRL) と通常の XBRL の両方に対応。

    iXBRL の場合、データは以下のタグに格納されている:
      <ix:nonFraction name="jppfs_cor:NetSales" contextRef="..." ...>123,456</ix:nonFraction>
      <ix:nonNumeric  name="jpdei_cor:FilerNameInJapaneseDEI" contextRef="...">会社名</ix:nonNumeric>

    Returns:
        list of dict: 各要素の情報（element, label_ja, value, context等）
    """
    is_ixbrl = filename.lower().endswith(('.htm', '.html'))

    if is_ixbrl:
        return _parse_ixbrl(content, filename)
    else:
        return _parse_regular_xbrl(content, filename)


def _get_all_text(elem):
    """要素内の全テキスト（子要素のテキスト含む）を取得"""
    return ''.join(elem.itertext()).strip()


def _parse_ixbrl(content: bytes, filename: str):
    """
    inline XBRL (iXBRL) を解析する。
    ix:nonFraction / ix:nonNumeric タグからデータを抽出。

    注意: lxml HTMLParser は属性名をすべて小文字にする。
      contextRef → contextref, unitRef → unitref 等
    """
    try:
        parser = etree.HTMLParser(encoding='utf-8')
        tree = etree.fromstring(content, parser)
    except Exception as e:
        print(f"   ❌ iXBRL解析エラー: {e}")
        return []

    # コンテキスト情報を取得（HTMLパーサー用: 属性名小文字対応）
    contexts = parse_contexts(tree)

    results = []

    # ix:nonFraction / ix:nonNumeric を探索
    # HTMLパーサーでは名前空間なしの "ix:nonfraction" / "ix:nonnumeric" として出現
    target_tags = {'ix:nonfraction', 'ix:nonnumeric'}

    for elem in tree.iter():
        tag = str(elem.tag).lower()

        if tag not in target_tags:
            continue

        # name 属性から要素名を取得 (例: "tse-ed-t:SalesIFRS", "jppfs_cor:NetSales")
        name_attr = elem.get("name", "")
        if not name_attr:
            continue

        # contextRef → HTMLパーサーで小文字化されて contextref
        context_ref = elem.get("contextref", "")
        if not context_ref:
            continue

        # 要素名を分解
        if ":" in name_attr:
            ns_prefix, element_name = name_attr.split(":", 1)
        else:
            ns_prefix = ""
            element_name = name_attr

        # テキスト値を取得（子要素内テキストも含む）
        text = _get_all_text(elem)

        # sign属性（HTMLパーサーは小文字化する）
        sign = elem.get("sign", "")
        # format属性
        fmt = elem.get("format", "")
        # scale属性（桁スケール: 例 scale="6" → 百万単位で表示された数値を円に変換）
        scale = elem.get("scale", "0")
        # unitref（小文字化）
        unit_ref = elem.get("unitref", "")
        # decimals
        decimals = elem.get("decimals", "")

        if not text:
            continue

        # 数値パース（ix:nonfraction の場合）
        value = None
        if tag == 'ix:nonfraction':
            try:
                clean = text.replace(",", "").replace("，", "").replace(" ", "").replace("\u3000", "")
                clean = clean.replace("△", "-").replace("▲", "-")
                if clean.startswith("(") and clean.endswith(")"):
                    clean = "-" + clean[1:-1]
                # ハイフン系（該当なし）はスキップ
                if clean in ("-", "－", "―", "—", ""):
                    continue
                value = float(clean)
                # sign属性
                if sign == "-":
                    value = -abs(value)
                # scale属性（スケーリング）
                try:
                    sc = int(scale)
                    if sc != 0:
                        value = value * (10 ** sc)
                except ValueError:
                    pass
            except (ValueError, TypeError):
                pass

        # 期間タイプの判定
        period_type = classify_period(context_ref, contexts)

        # TDnetサマリー要素名マッピング（tse-ed-t独自名 → 標準名）
        mapped_name = TSE_ELEMENT_MAP.get(element_name, element_name)
        label_ja = XBRL_LABEL_MAP.get(mapped_name, XBRL_LABEL_MAP.get(element_name, ""))

        results.append({
            "element": mapped_name,
            "label_ja": label_ja,
            "namespace": ns_prefix,
            "context_ref": context_ref,
            "period_type": period_type,
            "value": value,
            "value_raw": text,
            "unit_ref": unit_ref,
            "decimals": decimals,
        })

    return results


def _parse_regular_xbrl(content: bytes, filename: str):
    """通常の XBRL インスタンスを解析する。"""
    try:
        tree = etree.fromstring(content)
    except etree.XMLSyntaxError:
        try:
            parser = etree.HTMLParser()
            tree = etree.fromstring(content, parser)
        except Exception as e:
            print(f"   ❌ XBRL解析エラー: {e}")
            return []

    contexts = parse_contexts(tree)
    results = []

    for elem in tree.iter():
        tag = str(elem.tag)
        if '}' not in tag:
            continue

        local_name = etree.QName(tag).localname
        namespace = etree.QName(tag).namespace or ""

        context_ref = elem.get("contextRef")
        if context_ref is None:
            continue

        text = elem.text
        if text is None or text.strip() == "":
            continue
        text = text.strip()

        value = None
        try:
            clean = text.replace(",", "").replace("，", "")
            if clean.startswith("(") and clean.endswith(")"):
                clean = "-" + clean[1:-1]
            clean = clean.replace("△", "-").replace("▲", "-")
            value = float(clean)
        except (ValueError, TypeError):
            pass

        period_type = classify_period(context_ref, contexts)
        label_ja = XBRL_LABEL_MAP.get(local_name, "")

        ns_short = ""
        if namespace:
            if "jppfs" in namespace:
                ns_short = "jppfs_cor"
            elif "jpdei" in namespace:
                ns_short = "jpdei_cor"
            elif "jpcrp" in namespace:
                ns_short = "jpcrp_cor"
            elif "jpigp" in namespace:
                ns_short = "jpigp_cor"
            else:
                parts = namespace.rstrip("/").split("/")
                ns_short = parts[-1] if parts else namespace

        results.append({
            "element": local_name,
            "label_ja": label_ja,
            "namespace": ns_short,
            "context_ref": context_ref,
            "period_type": period_type,
            "value": value,
            "value_raw": text,
            "unit_ref": elem.get("unitRef", ""),
            "decimals": elem.get("decimals", ""),
        })

    return results


# ============================================================
# Section 3: DataFrame構築
# ============================================================

def build_dataframe(parsed_data: list) -> pd.DataFrame:
    """解析結果をDataFrameに変換"""
    df = pd.DataFrame(parsed_data)

    if df.empty:
        return df

    # ラベルが空のものにも要素名を表示
    df["display_name"] = df.apply(
        lambda row: row["label_ja"] if row["label_ja"] else row["element"],
        axis=1
    )

    return df


def build_financial_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    当期と前期のデータを横並びにした財務サマリーを構築する。
    同一要素で当期/前期の値を比較し、増減額・増減率を計算。
    """
    if df.empty:
        return pd.DataFrame()

    numeric_df = df[df["value"].notna()].copy()
    if numeric_df.empty:
        return pd.DataFrame()

    summary_rows = []
    elements = numeric_df["element"].unique()

    for elem_name in elements:
        elem_data = numeric_df[numeric_df["element"] == elem_name]
        label = XBRL_LABEL_MAP.get(elem_name, elem_name)

        current_val = None
        prior_val = None

        for _, row in elem_data.iterrows():
            pt = row["period_type"]
            val = row["value"]

            if pt in ("当期", "当期末", "当四半期"):
                if current_val is None:
                    current_val = val
            elif pt in ("前期", "前期末", "前四半期"):
                if prior_val is None:
                    prior_val = val

        if current_val is not None or prior_val is not None:
            # 増減額・増減率の計算
            change = None
            change_rate = None

            if current_val is not None and prior_val is not None and prior_val != 0:
                change = current_val - prior_val
                change_rate = change / abs(prior_val)

            summary_rows.append({
                "要素名": elem_name,
                "勘定科目": label,
                "当期": current_val,
                "前期": prior_val,
                "増減額": change,
                "増減率": change_rate,
            })

    return pd.DataFrame(summary_rows)


# ============================================================
# Section 4: 財務分析
# ============================================================

def analyze_significant_changes(summary_df: pd.DataFrame, threshold: float = 0.20) -> pd.DataFrame:
    """
    大きく増減変化した勘定科目を検出する。

    Args:
        summary_df: 財務サマリーDataFrame
        threshold:  変動率の閾値（デフォルト20%）

    Returns:
        DataFrame: 閾値以上の変動があった勘定科目（変動率の絶対値降順）
    """
    if summary_df.empty or "増減率" not in summary_df.columns:
        return pd.DataFrame()

    sig = summary_df[summary_df["増減率"].notna()].copy()
    sig = sig[sig["増減率"].abs() >= threshold]
    sig = sig.sort_values("増減率", ascending=False, key=abs)

    return sig


def calculate_profit_margins(summary_df: pd.DataFrame) -> pd.DataFrame:
    """
    各種利益率を計算する。

    計算する指標:
      - 売上総利益率   = 売上総利益 / 売上高
      - 営業利益率     = 営業利益 / 売上高
      - 経常利益率     = 経常利益 / 売上高
      - 当期純利益率   = 当期純利益 / 売上高
    """
    if summary_df.empty:
        return pd.DataFrame()

    # 売上高を取得（NetSales → Revenue → OperatingRevenue1 の順で探す）
    sales_current = None
    sales_prior = None

    for sales_elem in ["NetSales", "Revenue", "OperatingRevenue1"]:
        row = summary_df[summary_df["要素名"] == sales_elem]
        if not row.empty and row.iloc[0]["当期"] is not None:
            sales_current = row.iloc[0]["当期"]
            sales_prior = row.iloc[0]["前期"]
            break

    if sales_current is None or sales_current == 0:
        return pd.DataFrame()

    margin_items = [
        ("売上総利益率", "GrossProfit"),
        ("営業利益率", "OperatingIncome"),
        ("経常利益率", "OrdinaryIncome"),
        ("当期純利益率", "ProfitLoss"),
        ("親会社帰属純利益率", "ProfitLossAttributableToOwnersOfParent"),
    ]

    margin_rows = []
    for margin_name, elem_name in margin_items:
        row = summary_df[summary_df["要素名"] == elem_name]
        if row.empty:
            continue

        curr = row.iloc[0]["当期"]
        prev = row.iloc[0]["前期"]

        curr_margin = (curr / sales_current * 100) if curr is not None else None
        prev_margin = (prev / sales_prior * 100) if prev is not None and sales_prior and sales_prior != 0 else None

        diff = None
        if curr_margin is not None and prev_margin is not None:
            diff = curr_margin - prev_margin

        margin_rows.append({
            "指標": margin_name,
            "当期（%）": round(curr_margin, 2) if curr_margin is not None else None,
            "前期（%）": round(prev_margin, 2) if prev_margin is not None else None,
            "差分（pt）": round(diff, 2) if diff is not None else None,
        })

    return pd.DataFrame(margin_rows)


# ============================================================
# Section 5: Excel出力
# ============================================================

def export_to_excel(
    company_info: dict,
    summary_df: pd.DataFrame,
    significant_df: pd.DataFrame,
    margins_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    output_path: str,
):
    """分析結果を書式付きExcelファイルに出力"""

    wb = Workbook()

    # スタイル定義
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    alert_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    warn_fill = PatternFill(start_color="FFFFD0", end_color="FFFFD0", fill_type="solid")
    good_fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
    number_fmt = '#,##0'
    pct_fmt = '0.0%'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header_row(ws, row_num, num_cols):
        """ヘッダー行にスタイルを適用"""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    def auto_column_width(ws):
        """列幅を自動調整"""
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    text = str(cell.value)
                    length = len(text)
                    # 日本語文字は幅2倍扱い
                    for c in text:
                        if ord(c) > 127:
                            length += 1
                    max_length = max(max_length, length)
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    # ===================================================
    # Sheet 1: 分析サマリー
    # ===================================================
    ws1 = wb.active
    ws1.title = "分析サマリー"

    # 会社情報ヘッダー
    info_items = [
        ("会社名", company_info.get("name", "")),
        ("コード", company_info.get("code", "")),
        ("表題", company_info.get("title", "")),
        ("日付", company_info.get("date", "")),
    ]
    for i, (key, val) in enumerate(info_items, 1):
        ws1.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws1.cell(row=i, column=2, value=val)

    current_row = len(info_items) + 2

    # --- 利益率テーブル ---
    if not margins_df.empty:
        ws1.cell(row=current_row, column=1, value="【利益率分析】").font = Font(bold=True, size=12)
        current_row += 1

        for c_idx, col_name in enumerate(margins_df.columns, 1):
            ws1.cell(row=current_row, column=c_idx, value=col_name)
        style_header_row(ws1, current_row, len(margins_df.columns))
        current_row += 1

        for _, row in margins_df.iterrows():
            for c_idx, col_name in enumerate(margins_df.columns, 1):
                cell = ws1.cell(row=current_row, column=c_idx, value=row[col_name])
                cell.border = thin_border
                if col_name == "差分（pt）" and row[col_name] is not None:
                    if row[col_name] > 0:
                        cell.fill = good_fill
                    elif row[col_name] < -1:
                        cell.fill = alert_fill
            current_row += 1

        current_row += 1

    # --- 大幅変動テーブル ---
    if not significant_df.empty:
        ws1.cell(row=current_row, column=1, value="【大幅変動の勘定科目】").font = Font(bold=True, size=12)
        current_row += 1

        display_cols = ["勘定科目", "当期", "前期", "増減額", "増減率"]
        avail_cols = [c for c in display_cols if c in significant_df.columns]

        for c_idx, col_name in enumerate(avail_cols, 1):
            ws1.cell(row=current_row, column=c_idx, value=col_name)
        style_header_row(ws1, current_row, len(avail_cols))
        current_row += 1

        for _, row in significant_df.iterrows():
            for c_idx, col_name in enumerate(avail_cols, 1):
                val = row[col_name]
                cell = ws1.cell(row=current_row, column=c_idx, value=val)
                cell.border = thin_border

                if col_name in ("当期", "前期", "増減額") and isinstance(val, (int, float)):
                    cell.number_format = number_fmt
                elif col_name == "増減率" and isinstance(val, (int, float)):
                    cell.number_format = pct_fmt
                    if abs(val) >= 0.5:
                        cell.fill = alert_fill
                    elif abs(val) >= 0.3:
                        cell.fill = warn_fill
            current_row += 1

    auto_column_width(ws1)

    # ===================================================
    # Sheet 2: 財務データ一覧（当期/前期比較）
    # ===================================================
    if not summary_df.empty:
        ws2 = wb.create_sheet("財務データ一覧")

        display_cols = ["勘定科目", "当期", "前期", "増減額", "増減率"]
        avail_cols = [c for c in display_cols if c in summary_df.columns]

        for c_idx, col_name in enumerate(avail_cols, 1):
            ws2.cell(row=1, column=c_idx, value=col_name)
        style_header_row(ws2, 1, len(avail_cols))

        for r_idx, (_, row) in enumerate(summary_df.iterrows(), 2):
            for c_idx, col_name in enumerate(avail_cols, 1):
                val = row[col_name]
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border

                if col_name in ("当期", "前期", "増減額") and isinstance(val, (int, float)):
                    cell.number_format = number_fmt
                elif col_name == "増減率" and isinstance(val, (int, float)):
                    cell.number_format = pct_fmt
                    if abs(val) >= 0.3:
                        cell.fill = alert_fill
                    elif abs(val) >= 0.2:
                        cell.fill = warn_fill

        auto_column_width(ws2)

    # ===================================================
    # Sheet 3: XBRLデータ（Raw）
    # ===================================================
    if not raw_df.empty:
        ws3 = wb.create_sheet("XBRLデータ（Raw）")

        raw_display_cols = ["display_name", "element", "namespace", "period_type",
                            "value", "value_raw", "unit_ref", "context_ref"]
        avail_cols = [c for c in raw_display_cols if c in raw_df.columns]
        header_names = {
            "display_name": "勘定科目",
            "element": "XBRL要素名",
            "namespace": "名前空間",
            "period_type": "期間",
            "value": "数値",
            "value_raw": "原文",
            "unit_ref": "単位",
            "context_ref": "コンテキスト",
        }

        for c_idx, col_name in enumerate(avail_cols, 1):
            ws3.cell(row=1, column=c_idx, value=header_names.get(col_name, col_name))
        style_header_row(ws3, 1, len(avail_cols))

        for r_idx, (_, row) in enumerate(raw_df.iterrows(), 2):
            for c_idx, col_name in enumerate(avail_cols, 1):
                val = row[col_name]
                cell = ws3.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border
                if col_name == "value" and isinstance(val, (int, float)):
                    cell.number_format = number_fmt

        auto_column_width(ws3)

    # 保存
    wb.save(output_path)
    print(f"   📊 Excel出力: {output_path}")


# ============================================================
# 引数
# ============================================================

def parse_args():
    p = argparse.ArgumentParser(
        description="③ TDnet XBRL取得 & 財務分析ツール"
    )
    p.add_argument("--target", default=DEFAULT_TARGET_SPEC,
                    help="YYYYMMDD / YYYYMM / 'YYYYMMDD YYYYMMDD'")
    p.add_argument("--code", default=None,
                    help="証券コードでフィルタ（例: 7203）")
    p.add_argument("--save-root", default=DEFAULT_SAVE_ROOT,
                    help="保存先フォルダ")
    p.add_argument("--threshold", type=float, default=DEFAULT_CHANGE_THRESHOLD,
                    help="大幅変動の閾値（デフォルト0.20=20%%）")
    p.add_argument("--page-sleep", type=float, default=PAGE_SLEEP_SEC)
    p.add_argument("--xbrl-sleep", type=float, default=XBRL_SLEEP_SEC)
    return p.parse_args()


# ============================================================
# メイン処理
# ============================================================

def process_single_xbrl(zip_path, company_info, threshold, output_dir):
    """1つのXBRL ZIPファイルを解析・分析・Excel出力する"""

    # XBRL ZIPからインスタンスを探す
    filename, content = find_xbrl_instance_in_zip(zip_path)
    if content is None:
        print(f"   ⚠️ XBRLインスタンスが見つかりません: {zip_path}")
        return

    print(f"   📄 XBRLインスタンス: {filename}")

    # パース
    parsed_data = parse_xbrl_content(content, filename)
    if not parsed_data:
        print("   ⚠️ 財務データが抽出できませんでした")
        return

    print(f"   📊 抽出要素数: {len(parsed_data)}")

    # DataFrame構築
    raw_df = build_dataframe(parsed_data)
    summary_df = build_financial_summary(raw_df)

    # 分析
    significant_df = analyze_significant_changes(summary_df, threshold)
    margins_df = calculate_profit_margins(summary_df)

    # コンソールに結果表示
    if not margins_df.empty:
        print("\n   📈 【利益率】")
        for _, row in margins_df.iterrows():
            curr = f"{row['当期（%）']:.1f}%" if row['当期（%）'] is not None else "N/A"
            prev = f"{row['前期（%）']:.1f}%" if row['前期（%）'] is not None else "N/A"
            print(f"      {row['指標']}: 当期 {curr} ← 前期 {prev}")

    if not significant_df.empty:
        print(f"\n   ⚠️ 【大幅変動（閾値{threshold:.0%}以上）】")
        for _, row in significant_df.head(10).iterrows():
            label = row["勘定科目"] if row["勘定科目"] else row["要素名"]
            rate = row["増減率"]
            direction = "↑" if rate > 0 else "↓"
            print(f"      {direction} {label}: {rate:+.1%}")

    # Excel出力
    code = company_info.get("code", "unknown")
    name = safe_filename(company_info.get("name", "unknown"), max_len=20)
    excel_name = f"XBRL分析_{code}_{name}.xlsx"
    excel_path = output_dir / excel_name

    export_to_excel(company_info, summary_df, significant_df, margins_df, raw_df, str(excel_path))


def main():
    args = parse_args()

    global PAGE_SLEEP_SEC, XBRL_SLEEP_SEC
    PAGE_SLEEP_SEC = args.page_sleep
    XBRL_SLEEP_SEC = args.xbrl_sleep

    save_root = Path(args.save_root)
    save_root.mkdir(parents=True, exist_ok=True)

    d_from, d_to, label, mode = parse_target_spec(args.target)

    print("=" * 60)
    print("③ XBRL Financial Analyzer")
    print("=" * 60)
    print(f"🎯 対象期間: {d_from} ～ {d_to} (mode={mode})")
    if args.code:
        print(f"🔍 コードフィルタ: {args.code}")
    print(f"📁 保存先: {save_root}")
    print(f"📊 変動閾値: {args.threshold:.0%}")

    session = requests.Session()
    total_xbrl = 0
    total_analyzed = 0

    for target_date_str in iter_dates_yyyymmdd(d_from, d_to):
        print(f"\n{'=' * 60}")
        print(f"📅 日付: {target_date_str}")

        day_dir = save_root / target_date_str
        day_dir.mkdir(parents=True, exist_ok=True)

        # TDnetからXBRLリンクを取得
        xbrl_entries = find_xbrl_links(session, target_date_str, args.code)

        if not xbrl_entries:
            print("   📝 XBRLデータなし")
            continue

        print(f"   📦 XBRL対象: {len(xbrl_entries)} 件")

        for entry in xbrl_entries:
            code = entry["code"]
            name = entry["name"]
            title = entry["title"]

            print(f"\n   --- {code} {name} ---")
            print(f"   📄 {title}")

            # ZIPダウンロード
            zip_name = f"{safe_filename(code, 4)}_{safe_filename(name, 20)}_xbrl.zip"
            zip_path = day_dir / zip_name

            if zip_path.exists():
                print("   ⏭️ 既存ファイルあり（スキップ）")
            else:
                ok = download_xbrl_zip(session, entry["xbrl_url"], str(zip_path))
                if not ok:
                    continue
                print(f"   ✅ ダウンロード完了: {zip_name}")
                total_xbrl += 1

                if XBRL_SLEEP_SEC > 0:
                    time.sleep(XBRL_SLEEP_SEC)

            # 解析・分析
            try:
                company_info = {
                    "code": code,
                    "name": name,
                    "title": title,
                    "date": target_date_str,
                }
                process_single_xbrl(zip_path, company_info, args.threshold, day_dir)
                total_analyzed += 1
            except Exception as e:
                print(f"   ❌ 解析エラー: {e}")
                import traceback
                traceback.print_exc()

    print(f"\n{'=' * 60}")
    print("✅ ③完了")
    print(f"   XBRLダウンロード: {total_xbrl} 件")
    print(f"   分析完了: {total_analyzed} 件")
    print(f"   保存先: {save_root}")


if __name__ == "__main__":
    main()
