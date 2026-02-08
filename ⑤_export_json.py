# -*- coding: utf-8 -*-
"""
XBRL分析ExcelファイルをJSON形式にエクスポートする。
GitHub Pages公開用の index.json（一覧）と詳細JSONを生成。
yfinanceで株価指標（PBR, 予想PER, 配当利回り）も取得。

Usage:
  python "⑤_export_json.py"
  python "⑤_export_json.py" --force          # 全ファイル再生成
  python "⑤_export_json.py" --target 20260206 # 特定日付のみ
  python "⑤_export_json.py" --skip-stock     # 株価指標取得をスキップ
"""

import os, sys, re, json, math, argparse, time
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

DATA_ROOT = os.environ.get(
    "XBRL_DATA_ROOT",
    os.path.join(os.path.expanduser("~"), "Desktop", "XBRL_Data"),
)
DOCS_DIR = Path(__file__).parent / "docs"
DATA_DIR = DOCS_DIR / "data"
DETAIL_DIR = DATA_DIR / "detail"

# 売上高/営業収益として認識するラベル（日本語ラベルおよびXBRL要素名）
SALES_LABELS = {
    '売上高', '売上収益（IFRS）', '営業収益', '経常収益', '経常収益（保険）',
    'NetSales', 'NetSalesIFRS', 'NetSalesUS', 'TotalRevenuesUS',
    'Revenue', 'OperatingRevenue1',
    'OperatingRevenues', 'OperatingRevenuesIFRS',
    'OperatingRevenuesSpecific', 'OperatingRevenuesSE',
    'OrdinaryRevenuesBK', 'OrdinaryRevenuesIN',
}

# 売上高の増減率として認識するラベル
SALES_CHANGE_LABELS = {
    'ChangeInNetSales', 'ChangeInNetSalesIFRS',
    'ChangeInNetSalesUS', 'ChangeInTotalRevenuesUS',
    'ChangeInOperatingRevenues', 'ChangeInOperatingRevenuesIFRS',
    'ChangeInOperatingRevenuesSpecific', 'ChangeInOperatingRevenuesSE',
    'ChangeInOrdinaryRevenuesBK', 'ChangeInOrdinaryRevenuesIN',
}


def safe_val(v):
    """JSON互換の値に変換"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    return str(v)


def read_summary(p):
    """Excelからサマリー情報を読み取る（一覧用）"""
    from openpyxl import load_workbook
    info = {'title': '', 'op_cur': None, 'op_prev': None, 'op_diff': None, 'rev_chg': None}
    try:
        wb = load_workbook(str(p), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        info['title'] = str(ws.cell(row=3, column=2).value or '')

        # --- 営業利益率をSheet1から探す ---
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=4):
            if '営業利益率' in str(row[0].value or ''):
                for k, i in [('op_cur', 1), ('op_prev', 2), ('op_diff', 3)]:
                    v = row[i].value
                    info[k] = round(float(v), 2) if v is not None else None
                break

        # --- 財務データ一覧シートから売上高・営業利益を探す ---
        sales_cur = None
        sales_prev = None
        op_income_cur = None
        op_income_prev = None

        for sn in wb.sheetnames[1:]:
            ws2 = wb[sn]
            hdr = [str(c.value or '') for c in next(ws2.iter_rows(min_row=1, max_row=1))]

            # ヘッダーから列インデックスを特定
            col_map = {}
            for i, h in enumerate(hdr):
                if '当期' in h and '増減' not in h:
                    col_map['cur'] = i
                elif '前期' in h and '増減' not in h:
                    col_map['prev'] = i
                elif '増減率' in h:
                    col_map['rate'] = i

            if not col_map:
                continue

            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
                label = str(row[0].value or '').strip()

                # 売上高の増減率（増収率）
                if info['rev_chg'] is None:
                    if label in SALES_LABELS:
                        if 'rate' in col_map:
                            v = row[col_map['rate']].value
                            if v is not None:
                                info['rev_chg'] = round(float(v) * 100, 2)
                        # 売上高の当期/前期も記録（利益率計算用）
                        if 'cur' in col_map:
                            v = row[col_map['cur']].value
                            if v is not None:
                                sales_cur = float(v)
                        if 'prev' in col_map:
                            v = row[col_map['prev']].value
                            if v is not None:
                                sales_prev = float(v)
                    elif label in SALES_CHANGE_LABELS:
                        # 増減率が直接格納されている場合（当期列に率が入っている）
                        if 'cur' in col_map:
                            v = row[col_map['cur']].value
                            if v is not None and info['rev_chg'] is None:
                                info['rev_chg'] = round(float(v) * 100, 2)

                # 営業利益
                if label in ('営業利益', 'OperatingIncome'):
                    if 'cur' in col_map:
                        v = row[col_map['cur']].value
                        if v is not None:
                            op_income_cur = float(v)
                    if 'prev' in col_map:
                        v = row[col_map['prev']].value
                        if v is not None:
                            op_income_prev = float(v)

            # 見つかったら次のシートは不要
            if info['rev_chg'] is not None:
                break

        # --- 営業利益率がSheet1になかった場合、自力計算 ---
        if info['op_cur'] is None and sales_cur and sales_cur != 0 and op_income_cur is not None:
            info['op_cur'] = round(op_income_cur / sales_cur * 100, 2)
        if info['op_prev'] is None and sales_prev and sales_prev != 0 and op_income_prev is not None:
            info['op_prev'] = round(op_income_prev / sales_prev * 100, 2)
        if info['op_cur'] is not None and info['op_prev'] is not None and info['op_diff'] is None:
            info['op_diff'] = round(info['op_cur'] - info['op_prev'], 2)

        wb.close()
    except Exception as e:
        print(f"  Warning (summary): {e}")
    return info


def read_all_sheets(p):
    """Excelの全シートを読み取り、JSON互換のデータに変換"""
    from openpyxl import load_workbook
    wb = load_workbook(str(p), data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column, values_only=True):
            rows.append([safe_val(v) for v in row])
        sheets[name] = rows
    wb.close()
    return sheets


def fetch_stock_data(codes):
    """
    yfinanceで株価指標を一括取得する。
    codes: ユニークな証券コードのリスト
    Returns: {code: {pbr, forward_pe, trailing_pe, div_yield, price, market_cap}}
    """
    try:
        import yfinance as yf
    except ImportError:
        print("  Warning: yfinance未インストール。株価指標はスキップします。")
        print("  pip install yfinance でインストールしてください。")
        return {}

    result = {}
    total = len(codes)
    print(f"\n株価指標を取得中... ({total} 銘柄)")

    for i, code in enumerate(codes):
        if (i + 1) % 20 == 0 or i == 0:
            print(f"  [{i+1}/{total}] {code}.T ...")
        try:
            ticker = yf.Ticker(f"{code}.T")
            info = ticker.info

            def safe_round(v, n=2):
                if v is None:
                    return None
                try:
                    f = float(v)
                    return round(f, n) if not (math.isnan(f) or math.isinf(f)) else None
                except (ValueError, TypeError):
                    return None

            result[code] = {
                'pbr': safe_round(info.get('priceToBook')),
                'forward_pe': safe_round(info.get('forwardPE'), 1),
                'trailing_pe': safe_round(info.get('trailingPE'), 1),
                'div_yield': safe_round(info.get('dividendYield')),
                'price': safe_round(info.get('currentPrice'), 0),
                'market_cap': info.get('marketCap'),
            }
        except Exception as e:
            print(f"  Warning ({code}): {e}")
            result[code] = {}

        # レート制限対策（軽い間隔）
        if (i + 1) % 5 == 0:
            time.sleep(0.5)

    print(f"  取得完了: {len(result)} 銘柄")
    return result


def load_stock_cache():
    """株価キャッシュを読み込む（当日分のみ有効）"""
    cache_path = DATA_DIR / "stock_cache.json"
    if cache_path.exists():
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                cache = json.load(f)
            # 当日のキャッシュのみ有効
            import datetime
            today = datetime.date.today().strftime('%Y%m%d')
            if cache.get('_date') == today:
                print(f"  株価キャッシュ使用 ({today}, {len(cache) - 1} 銘柄)")
                return cache
        except Exception:
            pass
    return {}


def save_stock_cache(stock_data):
    """株価キャッシュを保存"""
    import datetime
    cache = dict(stock_data)
    cache['_date'] = datetime.date.today().strftime('%Y%m%d')
    cache_path = DATA_DIR / "stock_cache.json"
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def main():
    parser = argparse.ArgumentParser(description="Excel→JSON変換")
    parser.add_argument("--force", action="store_true", help="全ファイルを再生成")
    parser.add_argument("--target", type=str, help="特定日付のみ処理 (例: 20260206)")
    parser.add_argument("--skip-stock", action="store_true", help="株価指標取得をスキップ")
    args = parser.parse_args()

    root = Path(DATA_ROOT)
    if not root.exists():
        print(f"データディレクトリが見つかりません: {DATA_ROOT}")
        return

    DETAIL_DIR.mkdir(parents=True, exist_ok=True)

    index_entries = []
    generated = 0
    skipped = 0

    for dd in sorted(root.iterdir()):
        if not dd.is_dir() or not re.fullmatch(r'\d{8}', dd.name):
            continue
        d = dd.name
        if args.target and d != args.target:
            continue

        print(f"[{d}]")
        seen = {}

        # PDFリンク読み込み
        pdf_links = {}
        pdf_path = dd / "pdf_links.json"
        if pdf_path.exists():
            try:
                with open(pdf_path, 'r', encoding='utf-8') as f:
                    pdf_links = json.load(f)
            except Exception:
                pass

        for xf in sorted(dd.glob("XBRL*_*.xlsx")):
            m = re.match(r'XBRL[^_]*_([^_]+)_(.+)', xf.stem)
            if not m:
                continue
            code, company = m.group(1), m.group(2)

            # 詳細JSONファイル名（同一日付+コードの重複対応）
            base_key = f"{d}_{code}"
            if base_key in seen:
                seen[base_key] += 1
                detail_name = f"{base_key}_{seen[base_key]}.json"
            else:
                seen[base_key] = 0
                detail_name = f"{base_key}.json"

            # サマリー情報取得
            s = read_summary(xf)

            entry = {
                'date': f"{d[:4]}/{d[4:6]}/{d[6:]}",
                'date_raw': d,
                'code': code,
                'company': company,
                'title': s['title'],
                'rev_chg': safe_val(s['rev_chg']),
                'op_cur': safe_val(s['op_cur']),
                'op_prev': safe_val(s['op_prev']),
                'op_diff': safe_val(s['op_diff']),
                'detail': detail_name,
            }
            # PDFリンクがあれば追加
            if code in pdf_links:
                entry['pdf_url'] = pdf_links[code]
            index_entries.append(entry)

            # 詳細JSON生成（更新チェック）
            detail_path = DETAIL_DIR / detail_name
            excel_mtime = xf.stat().st_mtime
            need_update = args.force or not detail_path.exists()
            if not need_update and detail_path.exists():
                if detail_path.stat().st_mtime < excel_mtime:
                    need_update = True

            if need_update:
                try:
                    sheets = read_all_sheets(xf)
                    with open(detail_path, 'w', encoding='utf-8') as f:
                        json.dump({'sheets': sheets}, f, ensure_ascii=False)
                    print(f"  + {detail_name}")
                    generated += 1
                except Exception as e:
                    print(f"  ERROR {xf.name}: {e}")
            else:
                skipped += 1

    # --- 株価指標の取得 ---
    if not args.skip_stock and index_entries:
        # ユニークなコードを収集（4桁の数字またはアルファベット混在コード）
        unique_codes = sorted({e['code'] for e in index_entries
                               if re.fullmatch(r'[0-9A-Za-z]{4}', e['code'])})

        # キャッシュ確認（空データはリトライ対象）
        stock_cache = load_stock_cache()
        codes_to_fetch = []
        for c in unique_codes:
            cached = stock_cache.get(c)
            if cached is None:
                codes_to_fetch.append(c)
            elif isinstance(cached, dict) and not cached.get('pbr') and not cached.get('price'):
                # 前回取得失敗（空データ）→ リトライ
                codes_to_fetch.append(c)

        if codes_to_fetch:
            new_data = fetch_stock_data(codes_to_fetch)
            stock_cache.update(new_data)
            save_stock_cache(stock_cache)
        else:
            print(f"  株価データ: 全銘柄キャッシュ済み")

        # index_entries に株価データをマージ
        for entry in index_entries:
            sd = stock_cache.get(entry['code'], {})
            if isinstance(sd, dict) and '_date' not in sd:
                entry['pbr'] = sd.get('pbr')
                entry['forward_pe'] = sd.get('forward_pe')
                entry['div_yield'] = sd.get('div_yield')

    # index.json出力（常に再生成）
    index_path = DATA_DIR / "index.json"
    with open(index_path, 'w', encoding='utf-8') as f:
        json.dump(index_entries, f, ensure_ascii=False, indent=2)

    print(f"\n完了: {len(index_entries)}件")
    print(f"  詳細JSON: 生成={generated}, スキップ={skipped}")
    print(f"  出力先: {DATA_DIR}")


if __name__ == "__main__":
    main()
