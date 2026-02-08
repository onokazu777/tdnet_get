# -*- coding: utf-8 -*-
"""
XBRL Financial Viewer (Streamlit版)

Usage:  streamlit run ④_xbrl_viewer.py
Env:    XBRL_DATA_ROOT = データディレクトリ
"""

import os, re, json, math
from urllib.parse import quote, unquote
import streamlit as st
import pandas as pd
from pathlib import Path

# ============================================================
st.set_page_config(
    page_title="XBRL Financial Viewer",
    page_icon=":chart_with_upwards_trend:",
    layout="wide",
    initial_sidebar_state="collapsed",
)

DATA_ROOT = os.environ.get(
    "XBRL_DATA_ROOT",
    os.path.join(os.path.expanduser("~"), "Desktop", "XBRL_Data"),
)

# ============================================================
# CSS
# ============================================================
st.markdown("""<style>
.block-container{padding-top:.8rem;padding-bottom:1rem}
.section-hdr{font-size:.95rem;font-weight:700;margin:.6rem 0 .2rem}
.note-sm{font-size:.75rem;color:#888;margin-bottom:.2rem}
.xbrl-wrap{max-height:720px;overflow-y:auto;border-radius:4px;margin-bottom:.5rem}
.xbrl-table{width:100%;border-collapse:collapse;font-size:13px}
.xbrl-table thead th{position:sticky;top:0;background:#0f3460;color:#fff;padding:9px 12px;text-align:left;font-weight:600;border-bottom:2px solid #e94560;z-index:1;white-space:nowrap;cursor:default}
.xbrl-table thead th a{color:#fff;text-decoration:none;cursor:pointer}
.xbrl-table thead th a:hover{color:#e94560;text-decoration:underline}
.xbrl-table tbody tr{border-bottom:1px solid #2a2a4a;transition:background .12s}
.xbrl-table tbody tr:hover{background:#1e2a4a}
.xbrl-table tbody td{padding:7px 12px}
.xbrl-table a{color:#53b8f0;text-decoration:none;font-weight:500}
.xbrl-table a:hover{color:#e94560;text-decoration:underline}
.nm{text-align:right;font-variant-numeric:tabular-nums;font-family:Consolas,Menlo,monospace;font-size:12px}
.pos{color:#51cf66}.neg{color:#ff6b6b}.mu{color:#666}
</style>""", unsafe_allow_html=True)

# ============================================================
# 管理者認証
# ============================================================
if 'is_admin' not in st.session_state:
    st.session_state.is_admin = False

with st.sidebar:
    st.markdown("#### 管理者ログイン")
    _p = st.text_input("パスワード", type="password", label_visibility="collapsed",
                       placeholder="パスワード...")
    if _p:
        try:
            ok = _p == st.secrets.get("admin_password", "")
        except Exception:
            ok = False
        st.session_state.is_admin = ok
        if ok:
            st.success("認証OK")
        else:
            st.error("パスワードが違います")
    if st.session_state.is_admin:
        st.caption("Excel DL 有効")
        if st.button("ログアウト"):
            st.session_state.is_admin = False
            st.rerun()

# ============================================================
# 数値フォーマット
# ============================================================
def trunc(v, d=3):
    if not isinstance(v, (int, float)) or math.isnan(v) or math.isinf(v):
        return v
    f = 10**d
    return math.floor(v*f)/f if v >= 0 else math.ceil(v*f)/f

def fmt_amount(v):
    if not isinstance(v, (int, float)):
        return str(v) if v not in (None, "") else ""
    if math.isnan(v) or math.isinf(v): return ""
    if v == int(v) and abs(v) >= 1_000_000:
        return f"{int(v)//1_000_000:,}"
    t = trunc(v)
    return f"{int(t):,}" if t == int(t) else f"{t:,.3f}".rstrip('0').rstrip('.')

def fmt_rate(v):
    if not isinstance(v, (int, float)):
        return str(v) if v not in (None, "") else ""
    if math.isnan(v) or math.isinf(v): return ""
    return f"{trunc(v*100,2):.2f}%"

def fmt_pct(v):
    """利益率（%）・差分（pt）用: 小数点以下2位固定"""
    if not isinstance(v, (int, float)):
        return str(v) if v not in (None, "") else ""
    if math.isnan(v) or math.isinf(v): return ""
    return f"{trunc(v,2):.2f}"

def fmt_generic(v):
    if v is None or v == "": return ""
    if not isinstance(v, (int, float)): return str(v)
    if math.isnan(v) or math.isinf(v): return ""
    t = trunc(v)
    return f"{int(t):,}" if t == int(t) else f"{t:,.3f}".rstrip('0').rstrip('.')

def format_financial_df(df):
    r = df.copy()
    for col in r.columns:
        c = str(col)
        if '増減率' in c:
            r[col] = r[col].apply(lambda x: fmt_rate(x) if isinstance(x, (int, float)) else (str(x) if x else ""))
        elif '（%）' in c or '（pt）' in c:
            r[col] = r[col].apply(lambda x: fmt_pct(x) if isinstance(x, (int, float)) else (str(x) if x else ""))
        elif ('当期' in c or '前期' in c or '増減額' in c):
            r[col] = r[col].apply(lambda x: fmt_amount(x) if isinstance(x, (int, float)) else (str(x) if x else ""))
        else:
            r[col] = r[col].apply(lambda x: fmt_generic(x) if isinstance(x, (int, float)) else (str(x) if x is not None else ""))
    return r

def color_num(val):
    try:
        s = str(val).replace(',', '').replace('%', '').replace('pt', '').strip()
        if not s or s == '-': return ''
        v = float(s)
        if v < 0: return 'color: #ff6b6b'
        if v > 0: return 'color: #51cf66'
    except: pass
    return ''

# ============================================================
# データ読み込み
# ============================================================
def _read_summary(p):
    info = {'title': '', 'op_cur': None, 'op_prev': None, 'op_diff': None, 'rev_chg': None}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(p), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        info['title'] = str(ws.cell(row=3, column=2).value or '')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=4):
            if '営業利益率' in str(row[0].value or ''):
                for k, i in [('op_cur',1),('op_prev',2),('op_diff',3)]:
                    v = row[i].value
                    info[k] = round(float(v), 2) if v is not None else None
                break
        # 売上高の増減率（増収率）を財務データシートから取得
        SALES_LABELS = {'売上高', '売上収益（IFRS）', '営業収益'}
        for sn in wb.sheetnames[1:]:
            ws2 = wb[sn]
            hdr = [str(c.value or '') for c in next(ws2.iter_rows(min_row=1, max_row=1))]
            ri = None
            for i, h in enumerate(hdr):
                if '増減率' in h:
                    ri = i
                    break
            if ri is None:
                continue
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
                label = str(row[0].value or '').strip()
                if label in SALES_LABELS:
                    v = row[ri].value
                    if v is not None:
                        info['rev_chg'] = round(float(v) * 100, 2)
                    break
            if info['rev_chg'] is not None:
                break
        wb.close()
    except: pass
    return info

@st.cache_data(ttl=600, show_spinner="ファイルをスキャン中...")
def scan_files(data_root):
    root = Path(data_root)
    if not root.exists(): return []
    cp = root / "_index_cache.json"
    cache = {}
    if cp.exists():
        try:
            with open(cp, 'r', encoding='utf-8') as f: cache = json.load(f)
        except: cache = {}
    entries, upd = [], False
    for dd in sorted(root.iterdir()):
        if not dd.is_dir() or not re.fullmatch(r'\d{8}', dd.name): continue
        d = dd.name
        for xf in sorted(dd.glob("XBRL*_*.xlsx")):
            m = re.match(r'XBRL[^_]*_([^_]+)_(.+)', xf.stem)
            if not m: continue
            code, company = m.group(1), m.group(2)
            fk, mt = str(xf), xf.stat().st_mtime
            if fk in cache and cache[fk].get('mtime') == mt and 'rev_chg' in cache[fk]:
                c = cache[fk]
                title = c.get('title','')
                oc, op, od = c.get('op_cur'), c.get('op_prev'), c.get('op_diff')
                rc = c.get('rev_chg')
            else:
                s = _read_summary(xf)
                title, oc, op, od, rc = s['title'], s['op_cur'], s['op_prev'], s['op_diff'], s['rev_chg']
                cache[fk] = {'mtime':mt,'title':title,'op_cur':oc,'op_prev':op,'op_diff':od,'rev_chg':rc}
                upd = True
            entries.append({
                '日付': f"{d[:4]}/{d[4:6]}/{d[6:]}",
                'コード': code, '会社名': company,
                '表題': title.replace('[', '\\['),
                '増収率%': rc,
                '営利 当期%': oc, '営利 前期%': op, '営利 差分pt': od,
                '_path': fk, '_date': d,
            })
    if upd:
        try:
            with open(cp, 'w', encoding='utf-8') as f: json.dump(cache, f, ensure_ascii=False, indent=2)
        except: pass
    return entries

@st.cache_data(show_spinner="Excel読み込み中...")
def read_excel_detail(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    r = {}
    for n in wb.sheetnames:
        ws = wb[n]
        r[n] = [list(row) for row in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column, values_only=True)]
    wb.close()
    return r

# ============================================================
# 一覧テーブル（HTML）
# ============================================================
def _margin_html(val, unit):
    if val is None:
        return '<span class="mu">-</span>'
    cls = 'pos' if val > 0 else ('neg' if val < 0 else 'mu')
    return f'<span class="{cls}">{val:.2f}{unit}</span>'

# ソート可能カラム定義: 表示名 → (ソートキー, デフォルト昇順, 右寄せか)
_SORT_COLS = [
    ("日付",       "_date",      False, False),
    ("コード",     "コード",     True,  False),
    ("会社名",     None,         None,  False),   # ソート不可
    ("表題",       None,         None,  False),   # ソート不可
    ("増収率%",    "増収率%",    False, True),
    ("営利 当期%", "営利 当期%", False, True),
    ("営利 前期%", "営利 前期%", False, True),
    ("営利 差分pt","営利 差分pt",False, True),
]

def _sort_th(label, col_key, default_asc, right, cur_col, cur_asc):
    """ソート可能なヘッダーセルを生成"""
    align = ' style="text-align:right"' if right else ''
    if col_key is None:
        return f'<th{align}>{label}</th>'
    if cur_col == col_key:
        new_asc = not cur_asc
        arrow = ' ↑' if cur_asc else ' ↓'
    else:
        new_asc = default_asc
        arrow = ''
    href = f'?hsort={quote(col_key, safe="")}&hasc={"1" if new_asc else "0"}'
    return (f'<th{align}><a href="{href}" target="_self" '
            f'style="color:#fff;text-decoration:none">{label}{arrow}</a></th>')

def build_html_table(df, sort_col="_date", sort_asc=False):
    rows = ""
    for _, r in df.iterrows():
        pe = quote(r['_path'], safe='')
        title = str(r['表題']).replace('\\[', '[')
        rc = _margin_html(r['増収率%'], '%')
        oc = _margin_html(r['営利 当期%'], '%')
        op = _margin_html(r['営利 前期%'], '%')
        od = _margin_html(r['営利 差分pt'], 'pt')
        rows += (f'<tr>'
                 f'<td style="color:#8899aa;white-space:nowrap">{r["日付"]}</td>'
                 f'<td style="font-weight:600">{r["コード"]}</td>'
                 f'<td><a href="?view={pe}" target="_self">{r["会社名"]}</a></td>'
                 f'<td style="color:#b0b0c0;font-size:12px">{title}</td>'
                 f'<td class="nm">{rc}</td>'
                 f'<td class="nm">{oc}</td>'
                 f'<td class="nm">{op}</td>'
                 f'<td class="nm">{od}</td>'
                 f'</tr>')
    hdr = ''.join(_sort_th(lb, ck, da, rt, sort_col, sort_asc)
                  for lb, ck, da, rt in _SORT_COLS)
    return (f'<div class="xbrl-wrap"><table class="xbrl-table">'
            f'<thead><tr>{hdr}</tr></thead>'
            f'<tbody>{rows}</tbody></table></div>')

# ============================================================
# 分析サマリーパーサー
# ============================================================
def parse_summary_sections(rows):
    INFO_KEYS = {'会社名','コード','表題','日付'}
    ci, secs = {}, []
    cur_s, cur_h, cur_d = None, None, []
    for row in rows:
        f = str(row[0] or '').strip() if row and row[0] else ''
        if f in INFO_KEYS:
            ci[f] = row[1] if len(row) > 1 and row[1] else ''
        elif f.startswith('【'):
            if cur_s and cur_d: secs.append((cur_s, cur_h, cur_d))
            cur_s, cur_h, cur_d = f, None, []
        elif cur_s and cur_h is None and any(v for v in row if v is not None and v != ''):
            cur_h = [str(v or '') for v in row]
        elif cur_s and cur_h and any(v for v in row if v is not None and v != ''):
            cur_d.append(row)
    if cur_s and cur_d: secs.append((cur_s, cur_h, cur_d))
    return ci, secs

# ============================================================
# 一覧ページ
# ============================================================
def show_list(all_df):
    # ヘッダークリックによるソート処理
    hsort = st.query_params.get("hsort")
    if hsort:
        hasc = st.query_params.get("hasc", "0") == "1"
        st.session_state["_sort_col"] = unquote(hsort)
        st.session_state["_sort_asc"] = hasc
        st.query_params.clear()
        st.rerun()

    # ソート状態（セッションステートから読み取り、デフォルトは日付降順）
    sc = st.session_state.get("_sort_col", "_date")
    sa = st.session_state.get("_sort_asc", False)

    # セレクトボックスのソート（変更時はヘッダーソートをリセット）
    def _on_sort_select():
        st.session_state.pop("_sort_col", None)
        st.session_state.pop("_sort_asc", None)

    sm = {"日付 ↓新しい順":("_date",False),"日付 ↑古い順":("_date",True),
          "コード ↑昇順":("コード",True),"コード ↓降順":("コード",False),
          "増収率% ↓高い順":("増収率%",False),"増収率% ↑低い順":("増収率%",True),
          "営利 当期% ↓高い順":("営利 当期%",False),"営利 当期% ↑低い順":("営利 当期%",True),
          "営利 差分pt ↓高い順":("営利 差分pt",False),"営利 差分pt ↑低い順":("営利 差分pt",True)}

    c1, c2, c3 = st.columns([1.5, 2.5, 1])
    with c1:
        sl = st.selectbox("ソート", list(sm.keys()), label_visibility="collapsed",
                          key="_sort_select", on_change=_on_sort_select)
        # ヘッダーソートが無ければセレクトボックスの値を使う
        if "_sort_col" not in st.session_state:
            sc, sa = sm[sl]
    with c2:
        search = st.text_input("検索", placeholder="会社名・コード・表題で検索...", label_visibility="collapsed")
    with c3:
        dates = sorted(all_df['_date'].unique(), reverse=True)
        dl = ["全日付"] + [f"{d[:4]}/{d[4:6]}/{d[6:]}" for d in dates]
        dv = [""] + list(dates)
        di = st.selectbox("日付", range(len(dl)), format_func=lambda i: dl[i], label_visibility="collapsed")
        sd = dv[di]

    df = all_df.copy()
    if sd: df = df[df['_date'] == sd]
    if search:
        q = search.lower()
        df = df[df['会社名'].str.lower().str.contains(q, na=False) |
                df['コード'].str.lower().str.contains(q, na=False) |
                df['表題'].str.lower().str.contains(q, na=False)]
    df = df.sort_values(sc, ascending=sa, na_position='last')

    st.caption(f"{len(df)} / {len(all_df)} 件　— 会社名クリックで詳細 / ヘッダークリックでソート")
    st.markdown(build_html_table(df, sc, sa), unsafe_allow_html=True)

# ============================================================
# 詳細ダイアログ（一覧に重ねて表示）
# ============================================================
@st.dialog("詳細プレビュー", width="large")
def preview_dialog(item_dict, view_path):
    """一覧の上に重なるダイアログとして詳細を表示"""
    title_raw = str(item_dict.get('表題', '')).replace('\\[', '[')
    # HTMLエンティティに変換して Markdown リンク解釈を防止
    title_html = title_raw.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('[', '&#91;').replace(']', '&#93;')

    # ヘッダー行: 企業情報 + 閉じるボタン + DL
    hc, dc = st.columns([5, 1])
    with hc:
        st.markdown(f"**{item_dict['コード']}　{item_dict['会社名']}**")
        st.markdown(
            f"<span style='color:#888;font-size:0.85rem'>{item_dict['日付']}　|　{title_html}</span>",
            unsafe_allow_html=True)
    with dc:
        if st.button("✕ 閉じる", use_container_width=True):
            st.query_params.clear()
            st.rerun()
        if st.session_state.get('is_admin'):
            try:
                with open(view_path, 'rb') as f:
                    st.download_button("Excel DL", f.read(),
                                       file_name=os.path.basename(view_path),
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       use_container_width=True)
            except:
                pass

    st.divider()

    sheets = read_excel_detail(view_path)
    if not sheets:
        st.warning("データなし")
        return

    tabs = st.tabs(list(sheets.keys()))
    for tab, (name, rows) in zip(tabs, sheets.items()):
        with tab:
            if not rows:
                st.info("空のシート")
            elif name == "分析サマリー":
                _render_summary(rows)
            else:
                _render_data_sheet(rows)

def _render_summary(rows):
    ci, secs = parse_summary_sections(rows)
    if ci:
        cols = st.columns(min(len(ci), 4))
        for i, (k, v) in enumerate(ci.items()):
            v_safe = str(v).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('[','&#91;').replace(']','&#93;')
            cols[i % len(cols)].markdown(f"**{k}**<br>{v_safe}", unsafe_allow_html=True)
        st.divider()
    for sn, hd, dt in secs:
        st.markdown(f"<p class='section-hdr'>{sn}</p>", unsafe_allow_html=True)
        if not hd or not dt: continue
        # 空ヘッダー列を除去
        valid = [i for i, h in enumerate(hd) if h is not None and str(h).strip()]
        hd = [hd[i] for i in valid]
        dt = [[r[i] if i < len(r) else None for i in valid] for r in dt]
        mx = len(hd)
        padded = [list(r)[:mx] + [None]*max(0, mx-len(r)) for r in dt]
        df = pd.DataFrame(padded, columns=hd)
        hf = any('増減率' in str(h) or '当期' in str(h) or '前期' in str(h)
                 or '（%）' in str(h) or '（pt）' in str(h) for h in hd)
        if hf:
            df = format_financial_df(df)
            has_amount = any(('当期' in str(h) or '前期' in str(h)) and '（%）' not in str(h) and '（pt）' not in str(h) for h in hd)
            if has_amount:
                st.markdown("<p class='note-sm'>※ 金額は百万円単位 / 増減率は%表示</p>", unsafe_allow_html=True)
        nc = [c for c in df.columns if any(k in c for k in ['当期','前期','増減','差分'])]
        if nc:
            st.dataframe(df.style.map(color_num, subset=nc), hide_index=True, use_container_width=True)
        else:
            st.dataframe(df, hide_index=True, use_container_width=True)

def _render_data_sheet(rows):
    if len(rows) < 2:
        st.dataframe(pd.DataFrame(rows), hide_index=True)
        return
    hd = [str(v or f'列{i+1}') for i, v in enumerate(rows[0])]
    # 空ヘッダー列を除去
    valid = [i for i, h in enumerate(hd) if h.strip()]
    hd = [hd[i] for i in valid]
    mx = len(hd)
    data = []
    for r in rows[1:]:
        vals = [r[i] if i < len(r) else None for i in valid]
        if any(v is not None and v != '' for v in vals):
            data.append(vals[:mx] + [None]*max(0, mx-len(vals)))
    if not data:
        st.info("データなし"); return
    df = pd.DataFrame(data, columns=hd)
    hf = any('増減率' in str(h) or '当期' in str(h) or '前期' in str(h)
             or '（%）' in str(h) or '（pt）' in str(h) for h in hd)
    if hf:
        df = format_financial_df(df)
        has_amount = any(('当期' in str(h) or '前期' in str(h)) and '（%）' not in str(h) and '（pt）' not in str(h) for h in hd)
        if has_amount:
            st.markdown("<p class='note-sm'>※ 金額は百万円単位 / 増減率は%表示</p>", unsafe_allow_html=True)
    nc = [c for c in df.columns if any(k in c for k in ['当期','前期','増減','差分'])]
    if nc:
        st.dataframe(df.style.map(color_num, subset=nc), hide_index=True, use_container_width=True,
                     height=min(len(df)*35+60, 600))
    else:
        st.dataframe(df, hide_index=True, use_container_width=True,
                     height=min(len(df)*35+60, 600))

# ============================================================
# メイン
# ============================================================
def main():
    st.markdown("### XBRL Financial Viewer")
    entries = scan_files(DATA_ROOT)
    if not entries:
        st.warning(f"データが見つかりません: {DATA_ROOT}")
        return
    all_df = pd.DataFrame(entries)

    # 常に一覧を表示
    show_list(all_df)

    # クエリパラメータがあればダイアログを重ねて表示
    view_param = st.query_params.get("view")
    if view_param:
        view_path = unquote(view_param)
        info = all_df[all_df['_path'] == view_path]
        if not info.empty:
            preview_dialog(info.iloc[0].to_dict(), view_path)

if __name__ == "__main__":
    main()
